"""
Excel Import Tool for Finance Database
Parses a workbook sheet into up to 4 independent pools (cash, stocks, goods,
interest) and inserts each pool into its own dedicated DuckDB table. The
import entry point returns a structured dictionary contract so any pool can
be inspected/debugged in isolation, in addition to printing a readable
console summary.
"""

import calendar
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    from .db_schema import init_database, get_connection
except ImportError:
    from db_schema import init_database, get_connection


SUMMARY_TOKENS = ["bilanz", "summe", "gesamt", "subtotal", "total", "difference", "differenz", "kaufwert", "balance"]


class FinanceImporter:
    """Import a multi-pool Excel sheet (cash/stocks/goods/interest) into DuckDB."""

    def __init__(self, excel_path: str, db_path: str = "finance.duckdb"):
        self.excel_path = excel_path
        self.db_path = db_path
        self.conn = None
        self.xls = None
        self.sheet_name = None
        self.account_name = None
        self.transaction_date = None
        self.raw = None
        self.parsed_pools = {}

    # ------------------------------------------------------------------
    # Connection / sheet discovery
    # ------------------------------------------------------------------
    def connect(self):
        """Initialize database connection."""
        if self.db_path == ":memory:":
            # A fresh connection is required to keep the same in-memory database alive.
            self.conn = init_database(self.db_path)
            return

        if not Path(self.db_path).exists():
            init_database(self.db_path).close()
        self.conn = get_connection(self.db_path)

    def load_excel(self):
        """Load Excel file and show available sheets."""
        print(f"\n📂 Loading: {self.excel_path}")
        self.xls = pd.ExcelFile(self.excel_path)

        sheets = self.xls.sheet_names
        print(f"\n📋 Available sheets ({len(sheets)}):")
        for i, sheet in enumerate(sheets, 1):
            print(f"  {i}. {sheet}")

        return sheets

    def select_sheet(self, sheets):
        """Interactive sheet selection."""
        while True:
            try:
                choice = input(f"\nSelect sheet (1-{len(sheets)}) or name: ").strip()

                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(sheets):
                        self.sheet_name = sheets[idx]
                        break
                else:
                    if choice in sheets:
                        self.sheet_name = choice
                        break

                print("❌ Invalid selection")
            except KeyboardInterrupt:
                print("\n❌ Cancelled")
                return False

        print(f"✓ Selected: {self.sheet_name}")
        return True

    def handle_date_input(self):
        """Interactive date handling."""
        print("\n📅 Date Handling:")

        sheet_date = self._extract_date_from_sheet()

        if sheet_date:
            print(f"  Detected from sheet: {sheet_date.strftime('%B %Y')}")
            use_detected = input("  Use this date? (y/n): ").strip().lower()
            if use_detected == 'y':
                self.transaction_date = sheet_date
                return

        while True:
            date_str = input("  Enter transaction date (YYYY-MM-DD) or leave blank for today: ").strip()
            if not date_str:
                self.transaction_date = datetime.now().date()
                break
            try:
                self.transaction_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                break
            except ValueError:
                print("  ❌ Invalid format. Use YYYY-MM-DD")

        print(f"✓ Using date: {self.transaction_date}")

    def _extract_date_from_sheet(self):
        """Try to extract date from sheet name (e.g., 'August 2025' or 'Juni 2026')."""
        if not self.sheet_name or not isinstance(self.sheet_name, str):
            return None

        month_map = {
            "januar": 1,
            "jan": 1,
            "februar": 2,
            "feb": 2,
            "märz": 3,
            "maerz": 3,
            "marz": 3,
            "mar": 3,
            "april": 4,
            "apr": 4,
            "mai": 5,
            "juni": 6,
            "jun": 6,
            "juli": 7,
            "jul": 7,
            "august": 8,
            "aug": 8,
            "september": 9,
            "sep": 9,
            "oktober": 10,
            "okt": 10,
            "november": 11,
            "nov": 11,
            "dezember": 12,
            "dez": 12,
            "january": 1,
            "february": 2,
            "march": 3,
            "april": 4,
            "may": 5,
            "june": 6,
            "july": 7,
            "august": 8,
            "september": 9,
            "october": 10,
            "november": 11,
            "december": 12,
        }

        try:
            match = re.search(r'([A-Za-zÄÖÜäöüß]+)[\s\-_.]+(\d{4})', self.sheet_name)
            if match:
                month_str, year_str = match.groups()
                month_key = (
                    month_str.lower()
                    .replace("ä", "ae")
                    .replace("ö", "oe")
                    .replace("ü", "ue")
                    .replace("ß", "ss")
                )
                month = month_map.get(month_key)
                if month:
                    year = int(year_str)
                    last_day = calendar.monthrange(year, month)[1]
                    return date(year, month, last_day)
        except Exception:
            pass

        try:
            from dateutil.parser import parse
            match = re.search(r'([A-Za-zÄÖÜäöüß]+)[\s\-_.]+(\d{4})', self.sheet_name)
            if match:
                month_str, year_str = match.groups()
                date_str = f"{month_str} {year_str}"
                parsed = parse(date_str, default=datetime(1900, 1, 1))
                year = int(year_str)
                last_day = calendar.monthrange(year, parsed.month)[1]
                return date(year, parsed.month, last_day)
        except Exception:
            pass

        return None

    def _detect_account_name(self):
        """Detect account name from the cash table's account column.

        If the cash table contains explicit account labels in the last column,
        use the first account name found and keep it as a fallback for blank
        account rows. Otherwise fall back to merged header detection.
        """
        if self.raw is None:
            try:
                self.raw = self._load_raw_grid()
            except Exception:
                self.raw = None

        if self.raw is not None:
            cash_header = self._find_header_row(
                self.raw,
                [["nr.", "nr"], ["bezeichnung", "description"]],
                start=0,
            )
            if cash_header is not None:
                num_cols = self.raw.shape[1]
                account_col = num_cols - 1 if num_cols >= 2 else None
                first_account = None
                current_account = None

                for row_idx in range(cash_header + 1, len(self.raw)):
                    row = self.raw.iloc[row_idx]
                    if self._is_cash_summary_row(row, 0, 1):
                        break

                    if account_col is not None and account_col < len(row) and pd.notna(row.iat[account_col]):
                        raw_account = str(row.iat[account_col]).strip()
                        if raw_account and not raw_account.lower().startswith("unnamed"):
                            current_account = raw_account
                            if first_account is None:
                                first_account = current_account

                if first_account:
                    return first_account

        try:
            wb = load_workbook(self.excel_path, read_only=False, data_only=True)
            if self.sheet_name not in wb.sheetnames:
                return None
            ws = wb[self.sheet_name]
        except Exception:
            return None

        candidate = None
        try:
            merged_ranges = list(ws.merged_cells.ranges)
        except Exception:
            merged_ranges = []

        for merged_range in merged_ranges:
            min_row, max_row, min_col, max_col = (
                merged_range.min_row,
                merged_range.max_row,
                merged_range.min_col,
                merged_range.max_col,
            )
            if max_row <= 2:
                value = ws.cell(row=min_row, column=min_col).value
                if value and isinstance(value, str) and value.strip():
                    width = max_col - min_col + 1
                    if candidate is None or width > candidate["width"]:
                        candidate = {"value": value.strip(), "width": width}

        if candidate:
            return candidate["value"]

        first_row = [cell.value for cell in ws[1]]
        meaningful = [str(v).strip() for v in first_row if v and str(v).strip() and not str(v).strip().startswith("Unnamed")]
        if len(meaningful) == 1:
            return meaningful[0]

        return None

    # ------------------------------------------------------------------
    # Account helpers
    # ------------------------------------------------------------------
    def get_or_create_account(self, account_name):
        """Return account_id for an existing or newly created account."""
        if not account_name:
            account_name = self.account_name or "Imported Account"
        self.account_name = self.account_name or account_name

        existing_account = self.conn.execute(
            "SELECT account_id FROM accounts WHERE account_name = ?",
            [account_name]
        ).fetchone()
        if existing_account:
            return existing_account[0]

        max_id = self.conn.execute("SELECT MAX(account_id) FROM accounts").fetchone()[0] or 0
        next_id = max_id + 1
        self.conn.execute(
            "INSERT INTO accounts (account_id, account_name) VALUES (?, ?)",
            [next_id, account_name]
        )
        self.conn.commit()
        print(f"✓ Created new account: {account_name} (ID: {next_id})")
        return next_id

    # ------------------------------------------------------------------
    # Raw grid loading and section marker detection
    # ------------------------------------------------------------------
    def _load_raw_grid(self):
        return pd.read_excel(self.excel_path, sheet_name=self.sheet_name, header=None, engine="openpyxl")

    def _row_text(self, raw, row_idx):
        return " ".join(
            str(v).strip().lower()
            for v in raw.iloc[row_idx].tolist()
            if pd.notna(v)
        )

    def _find_marker_row(self, raw, tokens, start=0):
        for i in range(start, len(raw)):
            row_text = self._row_text(raw, i)
            if any(token in row_text for token in tokens):
                return i
        return None

    def _find_header_row(self, raw, tokens, require_tokens=None, start=0):
        def token_matches(token):
            if isinstance(token, (list, tuple, set)):
                return any(t in row_text for t in token)
            return token in row_text

        for i in range(start, len(raw)):
            row_text = self._row_text(raw, i)
            if all(token_matches(token) for token in tokens):
                if require_tokens is None or any(req in row_text for req in require_tokens):
                    return i
        return None

    def _find_marker_cell(self, raw, tokens, start=0):
        """Find the first (row, col) whose cell text contains any of the tokens."""
        for i in range(start, len(raw)):
            row = raw.iloc[i]
            for col_idx, value in enumerate(row.tolist()):
                if pd.isna(value):
                    continue
                text = str(value).strip().lower()
                if any(token in text for token in tokens):
                    return i, col_idx
        return None

    # ------------------------------------------------------------------
    # Amount / value parsing utilities
    # ------------------------------------------------------------------
    def _parse_amount(self, value):
        if pd.isna(value):
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if not text:
            return None

        text = text.replace("€", "").replace(" ", "")
        text = text.replace(".", "").replace(",", ".")
        text = text.replace("−", "-")

        try:
            return float(text)
        except (ValueError, TypeError):
            return None

    def _is_blank(self, value):
        return pd.isna(value) or (isinstance(value, str) and not value.strip())

    def _contains_summary_token(self, value):
        if not isinstance(value, str):
            return False
        lowered = value.lower()
        return any(token in lowered for token in SUMMARY_TOKENS)

    # ------------------------------------------------------------------
    # Pool detection
    # ------------------------------------------------------------------
    def _detect_pools(self):
        """Locate section boundaries for cash / stocks / goods / interest pools."""
        raw = self.raw
        total_rows = len(raw)

        pools = {
            "cash": None,
            "stocks": None,
            "goods": None,
            "interest": None,
        }

        cash_header = self._find_header_row(
            raw,
            [["nr.", "nr"], ["bezeichnung", "description"]],
            require_tokens=["einnahmen", "income", "bargeld", "cash"],
            start=0,
        )
        bilanz_row = self._find_marker_row(raw, ["bilanz:", "balance:"], start=(cash_header or 0) + 1)
        stock_marker = self._find_marker_row(raw, ["depot / wallet", "portfolio / wallet"], start=(bilanz_row or 0) + 1)
        goods_marker = self._find_marker_row(raw, ["sachwerte / abschreibung", "goods / depreciation"], start=(stock_marker or 0) + 1)

        if cash_header is not None:
            cash_end = total_rows
            if bilanz_row is not None:
                cash_end = min(cash_end, bilanz_row)
            if stock_marker is not None:
                cash_end = min(cash_end, stock_marker)
            pools["cash"] = {"header_row": cash_header, "end_row": cash_end - 1}

        if stock_marker is not None:
            stock_header = self._find_header_row(
                raw,
                [["nr.", "nr"], ["bezeichnung", "description"]],
                require_tokens=["wkn/symb", "ticker", "symbol", "kurs", "price"],
                start=stock_marker,
            )
            if stock_header is not None:
                stock_end = goods_marker if goods_marker is not None else total_rows
                pools["stocks"] = {"header_row": stock_header, "end_row": stock_end - 1}

        if goods_marker is not None:
            goods_header = self._find_header_row(
                raw,
                [["nr.", "nr"], ["bezeichnung", "description"]],
                require_tokens=["kaufwert", "purchase value", "abschreib", "depreciation"],
                start=goods_marker,
            )
            if goods_header is not None:
                pools["goods"] = {"header_row": goods_header, "end_row": total_rows - 1}

        # Interest / Festgeld / Sparen lives as a side-panel inside the stock section
        # (or anywhere after it), so scan the whole remainder of the sheet for the marker cell.
        interest_scan_start = stock_marker if stock_marker is not None else 0
        marker = self._find_marker_cell(raw, ["festgeld", "sparen", "fixed deposit", "interest", "savings"], start=interest_scan_start)
        if marker is not None:
            marker_row, marker_col = marker
            interest_end = goods_marker if goods_marker is not None else total_rows
            pools["interest"] = {"marker_row": marker_row, "marker_col": marker_col, "end_row": interest_end - 1}

        return pools

    # ------------------------------------------------------------------
    # Cash pool parsing
    # ------------------------------------------------------------------
    def _combine_header_levels(self, col):
        """Combine a pandas MultiIndex column tuple into one readable label.
        Relies on pandas' own merged-header forward-fill behavior for the
        top header level, so every column under a merged main category gets
        a consistent "Main / Sub" label.
        """
        parts = []
        for level in col:
            if pd.isna(level):
                continue
            text = str(level).strip()
            if text and not text.lower().startswith("unnamed"):
                parts.append(text)
        return " / ".join(parts) if parts else ""

    def _get_cash_column_labels(self, main_row, header_row, num_cols):
        if main_row == header_row:
            header_df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name, header=header_row, engine="openpyxl")
            labels = [
                "" if str(col).lower().startswith("unnamed") else str(col).strip()
                for col in header_df.columns
            ]
        else:
            header_df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name, header=[main_row, header_row], engine="openpyxl")
            labels = [self._combine_header_levels(col) for col in header_df.columns]

        if len(labels) < num_cols:
            labels += [""] * (num_cols - len(labels))
        return labels[:num_cols]

    def _parse_cash_pool(self, raw, header_row, end_row):
        """Parse the cash pool using structural column rules:
        col0=Nr, col1=Bezeichnung, middle cols=categories,
        second-to-last col=balance, last col=account.
        """
        if header_row is None or end_row <= header_row:
            return None

        num_cols = raw.shape[1]
        if num_cols < 4:
            return None

        main_row = header_row - 1 if header_row > 0 else header_row
        headers = self._get_cash_column_labels(main_row, header_row, num_cols)

        nr_col = 0
        desc_col = 1
        balance_col = num_cols - 2
        account_col = num_cols - 1
        category_cols = list(range(2, num_cols - 2))
        category_names = [headers[c] or f"Category {c}" for c in category_cols]

        records = []
        current_account = None
        for row_idx in range(header_row + 1, end_row + 1):
            row = raw.iloc[row_idx]

            if self._is_cash_summary_row(row, nr_col, desc_col):
                continue

            if account_col >= 0 and pd.notna(row.iat[account_col]):
                raw_account = str(row.iat[account_col]).strip()
                if raw_account and not raw_account.lower().startswith("unnamed"):
                    current_account = raw_account

            category_amounts = {}
            for col, cat_name in zip(category_cols, category_names):
                amount = self._parse_amount(row.iat[col])
                if amount:
                    category_amounts[cat_name] = category_amounts.get(cat_name, 0.0) + amount

            if not category_amounts:
                continue

            description = str(row.iat[desc_col]).strip() if pd.notna(row.iat[desc_col]) else "No description"
            row_nr = None
            if pd.notna(row.iat[nr_col]):
                try:
                    row_nr = int(float(row.iat[nr_col]))
                except (ValueError, TypeError):
                    row_nr = None

            balance = self._parse_amount(row.iat[balance_col]) if balance_col >= 0 else None
            account_name = current_account

            records.append({
                "row_nr": row_nr,
                "description": description,
                "account_name": account_name,
                "category_amounts": category_amounts,
                "balance_after": balance,
            })

        return {"records": records, "category_names": sorted(set(category_names))}

    def _is_cash_summary_row(self, row, nr_col, desc_col):
        description = str(row.iat[desc_col]).strip() if pd.notna(row.iat[desc_col]) else ""

        for value in row.tolist():
            if self._contains_summary_token(value):
                return True

        no_id = pd.isna(row.iat[nr_col])
        no_desc = description == ""
        if no_id and no_desc:
            numeric_values = [v for v in row if not pd.isna(v) and not isinstance(v, str)]
            if len(numeric_values) >= 3:
                return True

        return False

    # ------------------------------------------------------------------
    # Stock pool parsing
    # ------------------------------------------------------------------
    def _parse_stock_pool(self, raw, header_row, end_row):
        if header_row is None or end_row <= header_row:
            return None

        headers = [self._normalize_header_cell(v) for v in raw.iloc[header_row].tolist()]
        col_index = {name.lower(): idx for idx, name in enumerate(headers) if name}

        nr_col = self._match_col(col_index, ["nr.", "nr", "no."])
        name_col = self._match_col(col_index, ["bezeichnung", "description", "name"])
        ticker_col = self._match_col(col_index, ["wkn/symb", "ticker", "symbol"])
        qty_col = self._match_col(col_index, ["stück", "quantity", "qty", "shares"])
        price_prev_col = self._match_col(col_index, ["kurs alt", "price previous", "previous price", "price prev"])
        price_cur_col = self._match_col(col_index, ["kurs", "price", "current price"], exclude=["kurs alt"])
        delta_val_col = self._match_col(col_index, ["δ [€]", "delta [€]", "delta eur", "delta [eur]", "delta"])
        delta_pct_col = self._match_col(col_index, ["δ [%]", "delta [%]", "delta percent", "delta %"])
        position_value_col = self._match_col(
            col_index,
            ["gesamtwert", "total value", "position value", "market value", "value"],
        )
        depot_value_col = self._match_col(col_index, ["depotwert", "depot value", "portfolio value", "portfolio"])

        records = []
        for row_idx in range(header_row + 1, end_row + 1):
            row = raw.iloc[row_idx]

            row_nr = self._to_int(row.iat[nr_col]) if nr_col is not None else None
            name = str(row.iat[name_col]).strip() if name_col is not None and pd.notna(row.iat[name_col]) else ""

            # Structural validity rule: a real position row needs both a row number and a name.
            if row_nr is None or not name or self._contains_summary_token(name):
                continue

            records.append({
                "row_nr": row_nr,
                "name": name,
                "ticker": str(row.iat[ticker_col]).strip() if ticker_col is not None and pd.notna(row.iat[ticker_col]) else None,
                "quantity": self._parse_amount(row.iat[qty_col]) if qty_col is not None else None,
                "price_previous": self._parse_amount(row.iat[price_prev_col]) if price_prev_col is not None else None,
                "price_current": self._parse_amount(row.iat[price_cur_col]) if price_cur_col is not None else None,
                "delta_value": self._parse_amount(row.iat[delta_val_col]) if delta_val_col is not None else None,
                "delta_percent": self._parse_amount(row.iat[delta_pct_col]) if delta_pct_col is not None else None,
                "position_value": self._parse_amount(row.iat[position_value_col]) if position_value_col is not None else None,
                "depot_value_running": self._parse_amount(row.iat[depot_value_col]) if depot_value_col is not None else None,
            })

        return {"records": records}

    def _normalize_header_cell(self, value):
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.lower().startswith("unnamed"):
            return ""
        return text

    def _match_col(self, col_index, tokens, exclude=None):
        exclude = exclude or []
        for name_lower, idx in col_index.items():
            if any(ex in name_lower for ex in exclude):
                continue
            if any(token in name_lower for token in tokens):
                return idx
        return None

    def _to_int(self, value):
        if pd.isna(value):
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    # ------------------------------------------------------------------
    # Goods pool parsing
    # ------------------------------------------------------------------
    def _parse_goods_pool(self, raw, header_row, end_row):
        if header_row is None or end_row <= header_row:
            return None

        headers = [self._normalize_header_cell(v) for v in raw.iloc[header_row].tolist()]
        col_index = {name.lower(): idx for idx, name in enumerate(headers) if name}

        nr_col = self._match_col(col_index, ["nr.", "nr", "no."])
        name_col = self._match_col(col_index, ["bezeichnung", "description", "name", "item name"])
        purchase_col = self._match_col(col_index, ["kaufwert", "purchase value", "purchase"])
        depreciation_col = self._match_col(col_index, ["abschreib", "depreciation", "depreciation input"])
        value_prev_col = self._match_col(col_index, ["wert alt", "previous value", "value previous", "old value"])
        value_change_col = self._match_col(col_index, ["bilanz", "balance", "value change", "change", "delta"])
        current_value_col = self._match_col(
            col_index,
            ["wert", "value", "current value", "current_value"],
            exclude=["wert alt", "previous value", "purchase value", "purchase", "kaufwert"],
        )

        records = []
        for row_idx in range(header_row + 1, end_row + 1):
            row = raw.iloc[row_idx]

            row_nr = self._to_int(row.iat[nr_col]) if nr_col is not None else None
            name = str(row.iat[name_col]).strip() if name_col is not None and pd.notna(row.iat[name_col]) else ""

            if row_nr is None or not name or self._contains_summary_token(name):
                continue

            records.append({
                "row_nr": row_nr,
                "item_name": name,
                "purchase_value": self._parse_amount(row.iat[purchase_col]) if purchase_col is not None else None,
                "depreciation_input": self._parse_amount(row.iat[depreciation_col]) if depreciation_col is not None else None,
                "value_previous": self._parse_amount(row.iat[value_prev_col]) if value_prev_col is not None else None,
                "value_change": self._parse_amount(row.iat[value_change_col]) if value_change_col is not None else None,
                "current_value": self._parse_amount(row.iat[current_value_col]) if current_value_col is not None else None,
            })

        return {"records": records}

    # ------------------------------------------------------------------
    # Interest / Festgeld pool parsing (side-panel, label/value column pairs)
    # ------------------------------------------------------------------
    def _parse_interest_pool(self, raw, marker_row, marker_col, end_row):
        records = []
        scan_end = min(end_row + 1, len(raw)) if end_row is not None else len(raw)
        for row_idx in range(marker_row + 1, scan_end):
            row = raw.iloc[row_idx]
            label_value = row.iat[marker_col] if marker_col < len(row) else None

            if self._is_blank(label_value):
                break

            label_text = str(label_value).strip()
            if self._contains_summary_token(label_text):
                break

            balance = None
            for offset in range(1, 5):
                col = marker_col + offset
                if col >= len(row):
                    break
                candidate = self._parse_amount(row.iat[col])
                if candidate is not None:
                    balance = candidate
                    break

            if balance is None:
                continue

            records.append({
                "account_name": label_text,
                "balance": balance,
            })

        return {"records": records}

    # ------------------------------------------------------------------
    # Orchestration
    # ------------------------------------------------------------------
    def load_and_clean_data(self):
        """Load the raw grid and detect all present pools."""
        print(f"\n🔄 Loading data from '{self.sheet_name}'...")

        self.raw = self._load_raw_grid()
        if not self.account_name:
            detected = self._detect_account_name()
            if detected:
                self.account_name = detected
                print(f"✓ Detected account name from merged header: {self.account_name}")

        pools = self._detect_pools()

        self.parsed_pools = {}

        if pools["cash"]:
            self.parsed_pools["cash"] = self._parse_cash_pool(self.raw, pools["cash"]["header_row"], pools["cash"]["end_row"])
        if pools["stocks"]:
            self.parsed_pools["stocks"] = self._parse_stock_pool(self.raw, pools["stocks"]["header_row"], pools["stocks"]["end_row"])
        if pools["goods"]:
            self.parsed_pools["goods"] = self._parse_goods_pool(self.raw, pools["goods"]["header_row"], pools["goods"]["end_row"])
        if pools["interest"]:
            self.parsed_pools["interest"] = self._parse_interest_pool(
                self.raw, pools["interest"]["marker_row"], pools["interest"]["marker_col"], pools["interest"]["end_row"]
            )

        present = [name for name, data in self.parsed_pools.items() if data]
        print(f"✓ Detected pools: {', '.join(present) if present else 'none'}")

        return self.parsed_pools

    def ensure_categories_exist(self):
        """Create category entries for the cash pool if they don't already exist."""
        cash_pool = self.parsed_pools.get("cash")
        if not cash_pool:
            return

        print("\n📂 Ensuring categories exist...")
        for cat_name in cash_pool["category_names"]:
            existing = self.conn.execute(
                "SELECT category_id FROM categories WHERE category_name = ?",
                [cat_name]
            ).fetchone()

            if not existing:
                max_id = self.conn.execute("SELECT MAX(category_id) FROM categories").fetchone()[0] or 0
                next_id = max_id + 1

                lowered = str(cat_name).lower()
                income_keywords = [
                    "einnahmen",
                    "income",
                    "salary",
                    "lohn",
                    "gehalt",
                    "zins",
                    "interest",
                ]
                cat_type = 'income' if any(k in lowered for k in income_keywords) else 'expense'

                main_name, sub_name = self._split_category_name(cat_name)
                main_row = self.conn.execute(
                    "SELECT main_category_id FROM main_categories WHERE main_category_name = ?",
                    [main_name]
                ).fetchone()
                if main_row:
                    main_category_id = main_row[0]
                else:
                    main_category_id = self.conn.execute("SELECT COALESCE(MAX(main_category_id), 0) + 1 FROM main_categories").fetchone()[0]
                    self.conn.execute(
                        "INSERT INTO main_categories (main_category_id, main_category_name) VALUES (?, ?)",
                        [main_category_id, main_name]
                    )

                sub_category_id = None
                if sub_name is not None:
                    sub_row = self.conn.execute(
                        "SELECT sub_category_id FROM sub_categories WHERE main_category_id = ? AND sub_category_name = ?",
                        [main_category_id, sub_name]
                    ).fetchone()
                    if sub_row:
                        sub_category_id = sub_row[0]
                    else:
                        sub_category_id = self.conn.execute("SELECT COALESCE(MAX(sub_category_id), 0) + 1 FROM sub_categories").fetchone()[0]
                        self.conn.execute(
                            "INSERT INTO sub_categories (sub_category_id, main_category_id, sub_category_name) VALUES (?, ?, ?)",
                            [sub_category_id, main_category_id, sub_name]
                        )

                self.conn.execute(
                    "INSERT INTO categories (category_id, main_category_id, sub_category_id, category_name, category_type) VALUES (?, ?, ?, ?, ?)",
                    [next_id, main_category_id, sub_category_id, cat_name, cat_type]
                )
                print(f"  ✓ Created category: {cat_name}")

        self.conn.commit()

    def _split_category_name(self, category_name):
        if category_name is None:
            return "Uncategorized", None

        category_name = str(category_name).strip()
        if not category_name:
            return "Uncategorized", None

        if "/" in category_name:
            main_name, sub_name = category_name.split("/", 1)
            return main_name.strip() or "Uncategorized", sub_name

        return category_name, None

    def _insert_cash_pool(self, account_id):
        cash_pool = self.parsed_pools.get("cash")
        result = {"present": False, "rows_parsed": 0, "inserted": 0, "skipped": 0, "categories": []}
        if not cash_pool:
            return result

        result["present"] = True
        result["rows_parsed"] = len(cash_pool["records"])
        result["categories"] = cash_pool["category_names"]

        self.conn.execute("DELETE FROM cash_transactions WHERE source_sheet = ?", [self.sheet_name])

        running_txn = self.conn.execute("SELECT MAX(transaction_id) FROM cash_transactions").fetchone()[0] or 0
        next_txn_id = int(running_txn) + 1

        inserted = 0
        for record in cash_pool["records"]:
            row_account_name = record["account_name"] or self.account_name
            row_account_id = self.get_or_create_account(row_account_name) if row_account_name else account_id

            transaction_id = next_txn_id
            next_txn_id += 1

            for cat_name, amount in record["category_amounts"].items():
                cat_id = self.conn.execute(
                    "SELECT category_id FROM categories WHERE category_name = ?", [cat_name]
                ).fetchone()[0]

                self.conn.execute(
                    """INSERT OR REPLACE INTO cash_transactions
                    (transaction_id, row_nr, description, account_id, category_id, amount,
                     transaction_date, balance_after, source_sheet, section_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    [
                        transaction_id,
                        record["row_nr"],
                        record["description"],
                        row_account_id,
                        cat_id,
                        amount,
                        self.transaction_date,
                        record["balance_after"],
                        self.sheet_name,
                        "Cash",
                    ]
                )
                inserted += 1

        self.conn.commit()
        result["inserted"] = inserted
        result["skipped"] = result["rows_parsed"] - len([r for r in cash_pool["records"] if r["category_amounts"]])
        return result

    def _insert_stock_pool(self):
        stock_pool = self.parsed_pools.get("stocks")
        result = {"present": False, "rows_parsed": 0, "inserted": 0}
        if not stock_pool:
            return result

        result["present"] = True
        result["rows_parsed"] = len(stock_pool["records"])

        self.conn.execute("DELETE FROM stock_positions WHERE source_sheet = ?", [self.sheet_name])

        max_id = self.conn.execute("SELECT MAX(position_id) FROM stock_positions").fetchone()[0] or 0
        next_id = int(max_id) + 1

        for record in stock_pool["records"]:
            self.conn.execute(
                """INSERT INTO stock_positions
                (position_id, source_sheet, snapshot_date, row_nr, name, ticker, quantity,
                 price_previous, price_current, delta_value, delta_percent, position_value, depot_value_running)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    next_id,
                    self.sheet_name,
                    self.transaction_date,
                    record["row_nr"],
                    record["name"],
                    record["ticker"],
                    record["quantity"],
                    record["price_previous"],
                    record["price_current"],
                    record["delta_value"],
                    record["delta_percent"],
                    record["position_value"],
                    record["depot_value_running"],
                ]
            )
            next_id += 1

        self.conn.commit()
        result["inserted"] = len(stock_pool["records"])
        return result

    def _insert_goods_pool(self):
        goods_pool = self.parsed_pools.get("goods")
        result = {"present": False, "rows_parsed": 0, "inserted": 0}
        if not goods_pool:
            return result

        result["present"] = True
        result["rows_parsed"] = len(goods_pool["records"])

        self.conn.execute("DELETE FROM goods_valuations WHERE source_sheet = ?", [self.sheet_name])

        max_id = self.conn.execute("SELECT MAX(valuation_id) FROM goods_valuations").fetchone()[0] or 0
        next_id = int(max_id) + 1

        for record in goods_pool["records"]:
            self.conn.execute(
                """INSERT INTO goods_valuations
                (valuation_id, source_sheet, valuation_date, row_nr, item_name, purchase_value,
                 depreciation_input, value_previous, value_change, current_value)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    next_id,
                    self.sheet_name,
                    self.transaction_date,
                    record["row_nr"],
                    record["item_name"],
                    record["purchase_value"],
                    record["depreciation_input"],
                    record["value_previous"],
                    record["value_change"],
                    record["current_value"],
                ]
            )
            next_id += 1

        self.conn.commit()
        result["inserted"] = len(goods_pool["records"])
        return result

    def _insert_interest_pool(self):
        interest_pool = self.parsed_pools.get("interest")
        result = {"present": False, "rows_parsed": 0, "inserted": 0, "deltas": {}}
        if not interest_pool:
            return result

        result["present"] = True
        result["rows_parsed"] = len(interest_pool["records"])

        self.conn.execute("DELETE FROM interest_balance_changes WHERE source_sheet = ?", [self.sheet_name])

        max_id = self.conn.execute("SELECT MAX(change_id) FROM interest_balance_changes").fetchone()[0] or 0
        next_id = int(max_id) + 1

        for record in interest_pool["records"]:
            previous_row = self.conn.execute(
                """SELECT balance FROM interest_balance_changes
                   WHERE account_name = ? AND balance_date < ?
                   ORDER BY balance_date DESC LIMIT 1""",
                [record["account_name"], self.transaction_date]
            ).fetchone()
            previous_balance = float(previous_row[0]) if previous_row else None
            delta = (record["balance"] - previous_balance) if previous_balance is not None else None

            self.conn.execute(
                """INSERT INTO interest_balance_changes
                (change_id, source_sheet, balance_date, account_name, balance, previous_balance, delta)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                [
                    next_id,
                    self.sheet_name,
                    self.transaction_date,
                    record["account_name"],
                    record["balance"],
                    previous_balance,
                    delta,
                ]
            )
            result["deltas"][record["account_name"]] = delta
            next_id += 1

        self.conn.commit()
        result["inserted"] = len(interest_pool["records"])
        return result

    def import_sheet(self, sheet_name, account_name=None, transaction_date=None):
        """Run a non-interactive import workflow and return a dictionary result contract."""
        self.sheet_name = sheet_name
        self.account_name = account_name

        if transaction_date is None:
            self.transaction_date = self._extract_date_from_sheet() or datetime.now().date()
        elif isinstance(transaction_date, str):
            self.transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()
        else:
            self.transaction_date = transaction_date

        result = {
            "status": "success",
            "sheet_name": sheet_name,
            "account_name": account_name,
            "transaction_date": str(self.transaction_date),
            "pools": {},
            "warnings": [],
            "errors": [],
        }

        try:
            self.connect()
            self.load_and_clean_data()
            if result["account_name"] is None:
                result["account_name"] = self.account_name
            self.ensure_categories_exist()
            account_id = self.get_or_create_account(result["account_name"])

            result["pools"]["cash"] = self._insert_cash_pool(account_id)
            result["pools"]["stocks"] = self._insert_stock_pool()
            result["pools"]["goods"] = self._insert_goods_pool()
            result["pools"]["interest"] = self._insert_interest_pool()

            self._print_summary(result)
        except Exception as exc:  # noqa: BLE001
            result["status"] = "error"
            result["errors"].append(str(exc))
            raise

        return result

    def _print_summary(self, result):
        print("\n" + "=" * 60)
        print("IMPORT SUMMARY")
        print("=" * 60)
        print(f"📊 Account: {result['account_name']}")
        print(f"📅 Sheet: {result['sheet_name']}")
        print(f"📆 Date: {result['transaction_date']}\n")

        for pool_name, pool_result in result["pools"].items():
            if not pool_result.get("present"):
                print(f"  {pool_name.capitalize():<10} not present in sheet")
                continue
            print(f"  {pool_name.capitalize():<10} rows_parsed={pool_result.get('rows_parsed', 0)} inserted={pool_result.get('inserted', 0)}")

    def run(self):
        """Execute full interactive import workflow (CLI use)."""
        try:
            sheets = self.load_excel()

            if not self.select_sheet(sheets):
                return False

            self.handle_date_input()
            result = self.import_sheet(self.sheet_name, self.account_name, self.transaction_date)

            print("\n✅ Import completed successfully!")
            return result

        except Exception as e:
            print(f"\n❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            if self.conn:
                self.conn.close()


def main():
    """Main entry point."""
    excel_file = input("Enter path to Excel file: ").strip()

    if not Path(excel_file).exists():
        print(f"❌ File not found: {excel_file}")
        return

    importer = FinanceImporter(excel_file)
    importer.run()


if __name__ == "__main__":
    main()
